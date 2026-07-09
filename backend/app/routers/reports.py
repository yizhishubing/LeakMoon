"""
报表管理 API
作用：查询报表列表、生成Excel巡检报表
"""

import io
from datetime import datetime
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from app.database import get_db
from app.models.report import Report
from app.models.leak import LeakRecord
from app.models.website import Website
from app.models.alert import AlertLog
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

router = APIRouter()


# ===== Excel 样式 =====
HEADER_FONT = Font(name="微软雅黑", bold=True, size=12, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
TITLE_FONT = Font(name="微软雅黑", bold=True, size=14)
SUBTITLE_FONT = Font(name="微软雅黑", size=10, color="666666")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")


def _apply_header_style(ws, row, cols):
    """设置表头样式"""
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER


def _apply_cell_border(cell):
    """设置单元格边框"""
    cell.border = THIN_BORDER
    cell.alignment = Alignment(vertical="center", wrap_text=True)


@router.get("/")
def list_reports(db: Session = Depends(get_db)):
    """获取报表列表"""
    return db.query(Report).order_by(Report.generated_at.desc()).all()


@router.post("/")
def generate_report(data: dict, db: Session = Depends(get_db)):
    """
    生成巡检报表（Excel 文件）

    参数：
        data: {"title": str, "report_type": "daily/weekly/monthly/custom"}
    """
    title = data.get("title", f"巡检报表_{datetime.now().strftime('%Y%m%d_%H%M')}")
    report_type = data.get("report_type", "custom")

    # 统计数据
    total_leaks = db.query(func.count(LeakRecord.id)).scalar()
    high_count = db.query(func.count(LeakRecord.id)).filter(LeakRecord.severity == "high").scalar()
    medium_count = db.query(func.count(LeakRecord.id)).filter(LeakRecord.severity == "medium").scalar()
    low_count = db.query(func.count(LeakRecord.id)).filter(LeakRecord.severity == "low").scalar()

    # 告警统计
    total_alerts = db.query(func.count(AlertLog.id)).scalar()
    pending_alerts = db.query(func.count(AlertLog.id)).filter(
        AlertLog.status.in_(["pending", "sent"])
    ).scalar()

    summary = (
        f"本次巡检共发现 {total_leaks} 条泄露记录，"
        f"其中高风险 {high_count} 条，中风险 {medium_count} 条，低风险 {low_count} 条。"
        f"告警总数 {total_alerts} 条，待处理 {pending_alerts} 条。"
    )

    report = Report(
        title=title,
        report_type=report_type,
        summary=summary,
        status="generated",
    )
    db.add(report)
    db.commit()
    db.refresh(report)

    return {
        "id": report.id,
        "title": report.title,
        "report_type": report.report_type,
        "summary": report.summary,
        "generated_at": report.generated_at.isoformat(),
    }


@router.get("/download/{report_id}")
def download_report(report_id: int, db: Session = Depends(get_db)):
    """
    下载指定报表为 Excel 文件

    报表内容包含：
    1. 概览统计（总泄露数、分级统计、告警统计）
    2. 各类型泄露分布
    3. 各网站风险分布
    4. 最近泄露明细（前100条）
    """
    report = db.query(Report).filter(Report.id == report_id).first()
    if not report:
        raise HTTPException(status_code=404, detail="报表不存在")

    # 统计数据
    total_leaks = db.query(func.count(LeakRecord.id)).scalar()
    high_count = db.query(func.count(LeakRecord.id)).filter(LeakRecord.severity == "high").scalar()
    medium_count = db.query(func.count(LeakRecord.id)).filter(LeakRecord.severity == "medium").scalar()
    low_count = db.query(func.count(LeakRecord.id)).filter(LeakRecord.severity == "low").scalar()

    total_alerts = db.query(func.count(AlertLog.id)).scalar()
    pending_alerts = db.query(func.count(AlertLog.id)).filter(
        AlertLog.status.in_(["pending", "sent"])
    ).scalar()

    # 类型分布
    type_dist = (
        db.query(LeakRecord.data_type, func.count(LeakRecord.id))
        .group_by(LeakRecord.data_type)
        .all()
    )

    # 网站风险分布
    risk_dist = (
        db.query(
            Website.name,
            func.coalesce(func.sum(func.if_(LeakRecord.severity == "high", 1, 0)), 0).label("high"),
            func.coalesce(func.sum(func.if_(LeakRecord.severity == "medium", 1, 0)), 0).label("medium"),
            func.coalesce(func.sum(func.if_(LeakRecord.severity == "low", 1, 0)), 0).label("low"),
        )
        .outerjoin(LeakRecord, Website.id == LeakRecord.website_id)
        .group_by(Website.id, Website.name)
        .all()
    )

    # 最近泄露明细
    recent_leaks = (
        db.query(LeakRecord)
        .order_by(LeakRecord.detected_at.desc())
        .limit(100)
        .all()
    )

    # 生成 Excel
    wb = openpyxl.Workbook()

    # Sheet 1: 概览
    ws1 = wb.active
    ws1.title = "概览统计"
    ws1.sheet_properties.tabColor = "4472C4"

    ws1.merge_cells("A1:F1")
    cell = ws1["A1"]
    cell.value = report.title
    cell.font = Font(name="微软雅黑", bold=True, size=16, color="1F4E79")
    cell.alignment = CENTER_ALIGN

    ws1.merge_cells("A2:F2")
    cell = ws1["A2"]
    cell.value = f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}  |  报表类型：{report.report_type}"
    cell.font = SUBTITLE_FONT
    cell.alignment = CENTER_ALIGN

    # 统计卡片
    stats = [
        ("总泄露数", total_leaks),
        ("高风险", high_count),
        ("中风险", medium_count),
        ("低风险", low_count),
        ("告警总数", total_alerts),
        ("待处理", pending_alerts),
    ]

    row_start = 4
    ws1.cell(row=row_start, column=1, value="统计指标").font = HEADER_FONT
    ws1.cell(row=row_start, column=2, value="数值").font = HEADER_FONT
    for col in range(1, 3):
        ws1.cell(row=row_start, column=col).fill = HEADER_FILL
        ws1.cell(row=row_start, column=col).alignment = CENTER_ALIGN
        ws1.cell(row=row_start, column=col).border = THIN_BORDER

    for i, (label, value) in enumerate(stats):
        r = row_start + 1 + i
        ws1.cell(row=r, column=1, value=label).border = THIN_BORDER
        ws1.cell(row=r, column=2, value=value).border = THIN_BORDER
        ws1.cell(row=r, column=2).alignment = CENTER_ALIGN

    # Sheet 2: 类型分布
    ws2 = wb.create_sheet("类型分布")
    ws2.sheet_properties.tabColor = "70AD47"

    ws2["A1"].value = "泄露类型分布"
    ws2["A1"].font = TITLE_FONT

    ws2.cell(row=3, column=1, value="类型").font = HEADER_FONT
    ws2.cell(row=3, column=2, value="数量").font = HEADER_FONT
    for col in range(1, 3):
        ws2.cell(row=3, column=col).fill = HEADER_FILL
        ws2.cell(row=3, column=col).alignment = CENTER_ALIGN
        ws2.cell(row=3, column=col).border = THIN_BORDER

    for i, (dtype, count) in enumerate(type_dist):
        r = 4 + i
        ws2.cell(row=r, column=1, value=dtype).border = THIN_BORDER
        ws2.cell(row=r, column=2, value=count).border = THIN_BORDER
        ws2.cell(row=r, column=2).alignment = CENTER_ALIGN

    # Sheet 3: 网站风险
    ws3 = wb.create_sheet("网站风险")
    ws3.sheet_properties.tabColor = "FFC000"

    ws3["A1"].value = "各网站风险等级分布"
    ws3["A1"].font = TITLE_FONT

    headers = ["网站名称", "高风险", "中风险", "低风险", "合计"]
    for col, h in enumerate(headers, 1):
        ws3.cell(row=3, column=col, value=h).font = HEADER_FONT
        ws3.cell(row=3, column=col).fill = HEADER_FILL
        ws3.cell(row=3, column=col).alignment = CENTER_ALIGN
        ws3.cell(row=3, column=col).border = THIN_BORDER

    for i, (name, high, medium, low) in enumerate(risk_dist):
        r = 4 + i
        ws3.cell(row=r, column=1, value=name).border = THIN_BORDER
        ws3.cell(row=r, column=2, value=int(high)).border = THIN_BORDER
        ws3.cell(row=r, column=3, value=int(medium)).border = THIN_BORDER
        ws3.cell(row=r, column=4, value=int(low)).border = THIN_BORDER
        ws3.cell(row=r, column=5, value=int(high) + int(medium) + int(low)).border = THIN_BORDER
        for col in range(2, 6):
            ws3.cell(row=r, column=col).alignment = CENTER_ALIGN

    # Sheet 4: 泄露明细
    ws4 = wb.create_sheet("泄露明细")
    ws4.sheet_properties.tabColor = "C00000"

    ws4["A1"].value = f"最近泄露明细（共 {total_leaks} 条，展示前100条）"
    ws4["A1"].font = TITLE_FONT

    detail_headers = ["序号", "类型", "严重程度", "匹配内容", "来源URL", "检测时间"]
    for col, h in enumerate(detail_headers, 1):
        ws4.cell(row=3, column=col, value=h).font = HEADER_FONT
        ws4.cell(row=3, column=col).fill = HEADER_FILL
        ws4.cell(row=3, column=col).alignment = CENTER_ALIGN
        ws4.cell(row=3, column=col).border = THIN_BORDER

    for i, leak in enumerate(recent_leaks):
        r = 4 + i
        ws4.cell(row=r, column=1, value=i + 1).border = THIN_BORDER
        ws4.cell(row=r, column=1).alignment = CENTER_ALIGN
        ws4.cell(row=r, column=2, value=leak.data_type).border = THIN_BORDER
        ws4.cell(row=r, column=3, value=leak.severity).border = THIN_BORDER
        ws4.cell(row=r, column=3).alignment = CENTER_ALIGN
        ws4.cell(row=r, column=4, value=leak.matched_text).border = THIN_BORDER
        ws4.cell(row=r, column=5, value=leak.source_url).border = THIN_BORDER
        ws4.cell(row=r, column=6, value=leak.detected_at.strftime("%Y-%m-%d %H:%M") if leak.detected_at else "").border = THIN_BORDER

    # 设置列宽
    ws4.column_dimensions["A"].width = 6
    ws4.column_dimensions["B"].width = 15
    ws4.column_dimensions["C"].width = 12
    ws4.column_dimensions["D"].width = 30
    ws4.column_dimensions["E"].width = 60
    ws4.column_dimensions["F"].width = 18

    # 写入字节流
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # 设置响应头
    safe_name = "LeakMoon_Report"
    filename = f"{safe_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=report.xlsx"},
    )
